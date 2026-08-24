
from __future__ import annotations

import os
import shutil
import subprocess
import tempfile
from pathlib import Path

import openpyxl

from config import settings
from .excel_native_translator import ExcelNativeTranslator
from .excel_image_translator import ExcelImageTranslator


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _is_translatable_text(value) -> bool:
    if not isinstance(value, str):
        return False
    value = value.strip()
    if not value or _is_formula(value):
        return False
    if value.startswith(("http://", "https://", "mailto:")):
        return False
    return True


def _convert_xls_to_xlsx(path: Path) -> tuple[Path, bool]:
    """
    Legacy .xls conversion on Windows.

    Uses installed Microsoft Excel through COM when available, which is the
    safest route for preserving legacy workbook layout and embedded objects.
    Returns (xlsx_path, converted).
    """
    if path.suffix.lower() != ".xls":
        return path, False

    try:
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "Legacy .xls requires Microsoft Excel + pywin32 for layout/image-preserving "
            "conversion. Install pywin32 in the existing venv and ensure Excel is installed."
        ) from exc

    tmpdir = Path(tempfile.mkdtemp(prefix="excel_xls_"))
    output = tmpdir / f"{path.stem}.xlsx"

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    workbook = None
    try:
        workbook = excel.Workbooks.Open(str(path.resolve()), ReadOnly=True)
        # xlOpenXMLWorkbook = 51
        workbook.SaveAs(str(output), FileFormat=51)
    finally:
        if workbook is not None:
            workbook.Close(False)
        excel.Quit()

    return output, True


def translate_excel_workbook(
    input_path: str,
    output_path: str,
    api_key: str,
    model: str,
    image_model: str | None,
    source_language: str,
    target_language: str,
    translate_native_cells: bool = True,
    translate_images: bool = True,
):
    source = Path(input_path)
    requested_output = Path(output_path)

    converted_tmp = None
    workbook_path = source
    if source.suffix.lower() == ".xls":
        workbook_path, _ = _convert_xls_to_xlsx(source)
        converted_tmp = workbook_path.parent

    try:
        wb = openpyxl.load_workbook(
            workbook_path,
            data_only=False,
            keep_links=True,
        )

        native_cells = []
        if translate_native_cells:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if _is_translatable_text(cell.value):
                            native_cells.append((ws.title, cell.coordinate, cell.value))

        native_translations = {}
        if native_cells:
            translator = ExcelNativeTranslator(
                api_key=api_key,
                model=model,
                source_language=source_language,
                target_language=target_language,
                http_referer=settings.openrouter_http_referer,
                app_name=settings.openrouter_app_name,
            )
            native_translations = translator.translate_cells(native_cells)

            for (sheet, coordinate), translation in native_translations.items():
                wb[sheet][coordinate].value = translation

        image_results = []
        if translate_images:
            image_translator = ExcelImageTranslator(
                api_key=api_key,
                model=image_model or "",
                source_language=source_language,
                target_language=target_language,
                http_referer=settings.openrouter_http_referer,
                app_name=settings.openrouter_app_name,
            )
            image_results = image_translator.replace_sheet_images(wb)

        requested_output.parent.mkdir(parents=True, exist_ok=True)
        # Always save .xlsx. Legacy .xls is converted to .xlsx before processing.
        final_output = requested_output.with_suffix(".xlsx")
        wb.save(final_output)

        return {
            "output": str(final_output),
            "format": "xlsx",
            "sheets": wb.sheetnames,
            "native_cells_translated": len(native_translations),
            "native_cells_detected": len(native_cells),
            "images_translated": len(image_results),
            "image_results": [
                {
                    "sheet": x.sheet,
                    "index": x.index,
                    "model": x.model,
                    "original_display_size": list(x.original_size),
                    "generated_pixel_size": list(x.generated_size),
                }
                for x in image_results
            ],
        }
    finally:
        if converted_tmp:
            shutil.rmtree(converted_tmp, ignore_errors=True)
